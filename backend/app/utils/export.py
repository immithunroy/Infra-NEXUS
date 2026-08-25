"""Report export helpers (Excel + PDF).

Uses openpyxl for .xlsx and reportlab for .pdf. Both are optional at import
time so the API degrades gracefully if the extras are not installed.
"""
from __future__ import annotations

import io

from fastapi.responses import StreamingResponse


def _headers(title: str, filename: str) -> dict[str, str]:
    import urllib.parse
    from datetime import datetime, timezone

    disp = urllib.parse.quote(filename)
    return {
        "Content-Disposition": f"attachment; filename*=UTF-8''{disp}",
        "Cache-Control": "no-store",
    }


def xlsx_response(title: str, filename: str, headers: list[str], rows: list[list]):
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = (title or "Report")[:31]

    ws.append([title])
    ws["A1"].font = Font(bold=True, size=14)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(len(headers), 1))

    ws.append(headers)
    for c in range(1, len(headers) + 1):
        cell = ws.cell(row=2, column=c)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="334155")
        cell.alignment = Alignment(horizontal="center")

    for r in rows:
        ws.append(["" if v is None else v for v in r])

    # Auto-width
    for c in range(1, len(headers) + 1):
        letter = get_column_letter(c)
        vals = [headers[c - 1]] + [str(r[c - 1]) for r in rows if len(r) >= c]
        width = max((len(v) for v in vals), default=10)
        ws.column_dimensions[letter].width = min(max(width + 2, 10), 60)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=_headers(title, filename),
    )


def pdf_response(title: str, filename: str, headers: list[str], rows: list[list]):
    from datetime import datetime, timezone

    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
    from reportlab.lib.units import mm
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    from reportlab.platypus import PageBreak, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

    # Prefer DejaVuSans for wide Unicode coverage if present.
    font_name = "Helvetica"
    for path, name in [
        ("/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf", "DejaVuSans"),
        ("/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf", "DejaVuSans-Bold"),
    ]:
        try:
            import os

            if os.path.exists(path):
                pdfmetrics.registerFont(TTFont(name, path))
                font_name = "DejaVuSans"
        except Exception:
            pass

    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf,
        pagesize=landscape(A4),
        leftMargin=10 * mm,
        rightMargin=10 * mm,
        topMargin=12 * mm,
        bottomMargin=12 * mm,
        title=title,
    )
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "ReportTitle",
        parent=styles["Title"],
        fontName=font_name,
        fontSize=15,
        spaceAfter=4,
    )
    meta_style = ParagraphStyle(
        "Meta",
        parent=styles["Normal"],
        fontName=font_name,
        fontSize=8,
        textColor=colors.HexColor("#64748b"),
        spaceAfter=10,
    )
    cell_style = ParagraphStyle(
        "Cell",
        parent=styles["Normal"],
        fontName=font_name,
        fontSize=7.5,
        leading=9,
    )
    head_style = ParagraphStyle(
        "Head",
        parent=styles["Normal"],
        fontName=font_name,
        fontSize=7.5,
        leading=9,
        textColor=colors.white,
        alignment=1,
    )

    def clean(v) -> str:
        if v is None:
            return ""
        return str(v)

    story = [
        Paragraph(clean(title), title_style),
        Paragraph(
            f"Generated {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M UTC')}",
            meta_style,
        ),
    ]

    # Paginate rows into chunks (~46 rows/page landscape).
    page_rows = 46
    data_rows = [[Paragraph(clean(v), cell_style) for v in r] for r in rows]
    head_row = [Paragraph(clean(h), head_style) for h in headers]

    for start in range(0, max(len(data_rows), 1), page_rows):
        chunk = data_rows[start : start + page_rows]
        table = Table([head_row] + chunk, repeatRows=1, hAlign="LEFT")
        table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#334155")),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                    ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#cbd5e1")),
                    ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f8fafc")]),
                    ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                    ("LEFTPADDING", (0, 0), (-1, -1), 3),
                    ("RIGHTPADDING", (0, 0), (-1, -1), 3),
                    ("TOPPADDING", (0, 0), (-1, -1), 2),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 2),
                ]
            )
        )
        story.append(table)
        if start + page_rows < len(data_rows):
            story.append(PageBreak())

    doc.build(story)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/pdf",
        headers=_headers(title, filename),
    )