from fastapi import APIRouter, Depends, HTTPException, UploadFile, File
from fastapi.responses import StreamingResponse
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession
import io
import asyncio
import requests as http_requests

from ..database import get_db
from ..models import Cable, CableSegment, TjBox, Splitter, FiberLoop, CableCut, Splice, User
from ..schemas import (
    CableCreate, CableOut, CableUpdate,
    TjBoxCreate, TjBoxOut, TjBoxUpdate,
    SplitterCreate, SplitterOut, SplitterUpdate,
    FiberLoopCreate, FiberLoopOut, FiberLoopUpdate,
    CableCutCreate, CableCutOut, CableCutUpdate,
    SpliceCreate, SpliceOut, SpliceUpdate,
    CutRecoveryResult, CutRecoverySplice,
)
from ..security import get_current_user, require_write

router = APIRouter(prefix="/api/fiber", tags=["fiber"], dependencies=[Depends(get_current_user)])


# ------------------------------------------------------------------ Cables
@router.get("/cables", response_model=list[CableOut])
async def list_cables(db: AsyncSession = Depends(get_db)):
    res = await db.execute(select(Cable).order_by(Cable.id))
    cables = res.scalars().all()
    result = []
    for c in cables:
        seg_res = await db.execute(select(CableSegment).where(CableSegment.cable_id == c.id).order_by(CableSegment.order_index))
        c.segments = seg_res.scalars().all()
        c.src_tj_name = ""
        c.dst_tj_name = ""
        if c.src_tj_id:
            tj = await db.get(TjBox, c.src_tj_id)
            if tj: c.src_tj_name = tj.name
        if c.dst_tj_id:
            tj = await db.get(TjBox, c.dst_tj_id)
            if tj: c.dst_tj_name = tj.name
        result.append(c)
    return result


@router.post("/cables", response_model=CableOut)
async def create_cable(body: CableCreate, user: User = Depends(require_write), db: AsyncSession = Depends(get_db)):
    # Auto-generate link_id
    max_res = await db.execute(select(Cable.id).order_by(Cable.id.desc()).limit(1))
    max_id = max_res.scalar() or 0
    link_id = f"LINK-{max_id + 1001}"

    cable = Cable(
        link_id=link_id, link_name=body.link_name, code=body.code,
        core_count=body.core_count, manufacturer=body.manufacturer,
        manufacturing_year=body.manufacturing_year, cable_type=body.cable_type,
        route_type=body.route_type, src_tj_id=body.src_tj_id, dst_tj_id=body.dst_tj_id,
        notes=body.notes,
    )
    db.add(cable)
    await db.flush()

    segments = list(body.segments)
    if segments:
        for i, seg in enumerate(segments):
            db.add(CableSegment(
                cable_id=cable.id,
                start_lat=seg.start_lat, start_lng=seg.start_lng,
                end_lat=seg.end_lat, end_lng=seg.end_lng,
                order_index=seg.order_index if seg.order_index is not None else i,
            ))
    elif cable.src_tj_id and cable.dst_tj_id:
        src_tj = await db.get(TjBox, cable.src_tj_id)
        dst_tj = await db.get(TjBox, cable.dst_tj_id)
        if src_tj and dst_tj:
            profile = cable.route_type or "driving"
            coords = f"{src_tj.lng},{src_tj.lat};{dst_tj.lng},{dst_tj.lat}"
            url = f"https://router.project-osrm.org/route/v1/{profile}/{coords}?overview=full&geometries=geojson"
            try:
                def _fetch():
                    return http_requests.get(url, timeout=10)
                resp = await asyncio.to_thread(_fetch)
                data = resp.json()
                if data.get("code") == "Ok" and data["routes"]:
                    coords_list = data["routes"][0]["geometry"]["coordinates"]
                    for i in range(len(coords_list) - 1):
                        db.add(CableSegment(
                            cable_id=cable.id,
                            start_lat=coords_list[i][1], start_lng=coords_list[i][0],
                            end_lat=coords_list[i+1][1], end_lng=coords_list[i+1][0],
                            order_index=i,
                        ))
            except Exception:
                pass

    await db.commit()
    await db.refresh(cable)
    seg_res = await db.execute(select(CableSegment).where(CableSegment.cable_id == cable.id).order_by(CableSegment.order_index))
    cable.segments = seg_res.scalars().all()
    return cable


@router.put("/cables/{cable_id}", response_model=CableOut)
async def update_cable(cable_id: int, body: CableUpdate, user: User = Depends(require_write), db: AsyncSession = Depends(get_db)):
    cable = await db.get(Cable, cable_id)
    if cable is None:
        raise HTTPException(status_code=404, detail="Cable not found")
    for field, value in body.model_dump(exclude_unset=True, exclude={"segments"}).items():
        setattr(cable, field, value)
    if body.segments is not None:
        from sqlalchemy import delete
        await db.execute(delete(CableSegment).where(CableSegment.cable_id == cable_id))
        for i, seg in enumerate(body.segments):
            db.add(CableSegment(
                cable_id=cable_id, start_lat=seg.start_lat, start_lng=seg.start_lng,
                end_lat=seg.end_lat, end_lng=seg.end_lng, order_index=i,
            ))
    await db.commit()
    await db.refresh(cable)
    seg_res = await db.execute(select(CableSegment).where(CableSegment.cable_id == cable.id).order_by(CableSegment.order_index))
    cable.segments = seg_res.scalars().all()
    return cable


@router.delete("/cables/{cable_id}", status_code=204)
async def delete_cable(cable_id: int, user: User = Depends(require_write), db: AsyncSession = Depends(get_db)):
    cable = await db.get(Cable, cable_id)
    if cable is None:
        raise HTTPException(status_code=404, detail="Cable not found")
    await db.delete(cable)
    await db.commit()


# ------------------------------------------------------------------ TJ Boxes
@router.get("/tj-boxes", response_model=list[TjBoxOut])
async def list_tj_boxes(db: AsyncSession = Depends(get_db)):
    res = await db.execute(select(TjBox).order_by(TjBox.id))
    return res.scalars().all()


@router.post("/tj-boxes", response_model=TjBoxOut)
async def create_tj_box(body: TjBoxCreate, user: User = Depends(require_write), db: AsyncSession = Depends(get_db)):
    res = await db.execute(select(TjBox.unique_id).where(TjBox.unique_id.like("TJ-%")).order_by(TjBox.id.desc()).limit(1))
    last = res.scalar()
    next_num = 5001
    if last:
        try:
            next_num = int(last.split("-")[1]) + 1
        except (IndexError, ValueError):
            pass
    data = body.model_dump()
    data["capacity"] = data.get("tray_count", 1) * data.get("splice_per_tray", 12)
    box = TjBox(unique_id=f"TJ-{next_num}", **data)
    db.add(box)
    await db.commit()
    await db.refresh(box)
    return box


@router.put("/tj-boxes/{box_id}", response_model=TjBoxOut)
async def update_tj_box(box_id: int, body: TjBoxUpdate, user: User = Depends(require_write), db: AsyncSession = Depends(get_db)):
    box = await db.get(TjBox, box_id)
    if box is None:
        raise HTTPException(status_code=404, detail="TJ Box not found")
    for field, value in body.model_dump(exclude_unset=True).items():
        setattr(box, field, value)
    box.capacity = box.tray_count * box.splice_per_tray
    await db.commit()
    await db.refresh(box)
    return box


@router.put("/tj-boxes/{box_id}/move")
async def move_tj_box(box_id: int, body: TjBoxUpdate, user: User = Depends(require_write), db: AsyncSession = Depends(get_db)):
    """Move a TJ box to new coordinates. Updates connected cable endpoints automatically."""
    from sqlalchemy import or_
    
    box = await db.get(TjBox, box_id)
    if box is None:
        raise HTTPException(status_code=404, detail="TJ Box not found")
    
    new_lat = body.lat if body.lat is not None else box.lat
    new_lng = body.lng if body.lng is not None else box.lng
    
    # Calculate offset
    lat_offset = new_lat - box.lat
    lng_offset = new_lng - box.lng
    
    if lat_offset == 0 and lng_offset == 0:
        return box
    
    # Update TJ location
    box.lat = new_lat
    box.lng = new_lng
    await db.flush()
    
    # Update connected cable segments - move endpoints that are near the old TJ location
    cables_res = await db.execute(
        select(Cable).where(or_(Cable.src_tj_id == box_id, Cable.dst_tj_id == box_id))
    )
    cables = cables_res.scalars().all()
    
    for cable in cables:
        segs_res = await db.execute(
            select(CableSegment).where(CableSegment.cable_id == cable.id).order_by(CableSegment.order_index)
        )
        segments = segs_res.scalars().all()
        
        if not segments:
            continue
        
        # Update first segment start if src TJ
        if cable.src_tj_id == box_id:
            segments[0].start_lat = new_lat
            segments[0].start_lng = new_lng
        
        # Update last segment end if dst TJ
        if cable.dst_tj_id == box_id:
            segments[-1].end_lat = new_lat
            segments[-1].end_lng = new_lng
        
        # Update intermediate segment endpoints if they match old position
        for seg in segments:
            if abs(seg.start_lat - (new_lat - lat_offset)) < 0.0001 and abs(seg.start_lng - (new_lng - lng_offset)) < 0.0001:
                seg.start_lat = new_lat
                seg.start_lng = new_lng
            if abs(seg.end_lat - (new_lat - lat_offset)) < 0.0001 and abs(seg.end_lng - (new_lng - lng_offset)) < 0.0001:
                seg.end_lat = new_lat
                seg.end_lng = new_lng
    
    # Update hosted splitters
    splitters_res = await db.execute(select(Splitter).where(Splitter.tj_box_id == box_id))
    splitters = splitters_res.scalars().all()
    for splitter in splitters:
        splitter.lat = new_lat
        splitter.lng = new_lng
    
    await db.commit()
    await db.refresh(box)
    return box


@router.delete("/tj-boxes/{box_id}", status_code=204)
async def delete_tj_box(box_id: int, user: User = Depends(require_write), db: AsyncSession = Depends(get_db)):
    box = await db.get(TjBox, box_id)
    if box is None:
        raise HTTPException(status_code=404, detail="TJ Box not found")
    # Delete hosted splitters first
    splitters = (await db.execute(select(Splitter).where(Splitter.tj_box_id == box_id))).scalars().all()
    for sp in splitters:
        await db.delete(sp)
    await db.delete(box)
    await db.commit()


# ------------------------------------------------------------------ Splitters
@router.get("/splitters", response_model=list[SplitterOut])
async def list_splitters(db: AsyncSession = Depends(get_db)):
    res = await db.execute(select(Splitter).order_by(Splitter.id))
    splitters = res.scalars().all()
    result = []
    for s in splitters:
        if s.tj_box_id:
            box = await db.get(TjBox, s.tj_box_id)
            s.tj_box_name = box.name if box else ""
        result.append(s)
    return result


@router.post("/splitters", response_model=SplitterOut)
async def create_splitter(body: SplitterCreate, user: User = Depends(require_write), db: AsyncSession = Depends(get_db)):
    res = await db.execute(select(Splitter.unique_id).where(Splitter.unique_id.like("SP-%")).order_by(Splitter.id.desc()).limit(1))
    last = res.scalar()
    next_num = 1001
    if last:
        try:
            next_num = int(last.split("-")[1]) + 1
        except (IndexError, ValueError):
            pass
    splitter = Splitter(unique_id=f"SP-{next_num}", **body.model_dump())
    db.add(splitter)
    await db.commit()
    await db.refresh(splitter)
    return splitter


@router.put("/splitters/{splitter_id}", response_model=SplitterOut)
async def update_splitter(splitter_id: int, body: SplitterUpdate, user: User = Depends(require_write), db: AsyncSession = Depends(get_db)):
    splitter = await db.get(Splitter, splitter_id)
    if splitter is None:
        raise HTTPException(status_code=404, detail="Splitter not found")
    for field, value in body.model_dump(exclude_unset=True).items():
        setattr(splitter, field, value)
    await db.commit()
    await db.refresh(splitter)
    return splitter


@router.delete("/splitters/{splitter_id}", status_code=204)
async def delete_splitter(splitter_id: int, user: User = Depends(require_write), db: AsyncSession = Depends(get_db)):
    splitter = await db.get(Splitter, splitter_id)
    if splitter is None:
        raise HTTPException(status_code=404, detail="Splitter not found")
    await db.delete(splitter)
    await db.commit()


# ------------------------------------------------------------------ Fiber Loops
@router.get("/loops", response_model=list[FiberLoopOut])
async def list_loops(cable_id: int | None = None, db: AsyncSession = Depends(get_db)):
    q = select(FiberLoop).order_by(FiberLoop.id)
    if cable_id is not None:
        q = q.where(FiberLoop.cable_id == cable_id)
    res = await db.execute(q)
    return res.scalars().all()


@router.post("/loops", response_model=FiberLoopOut)
async def create_loop(body: FiberLoopCreate, user: User = Depends(require_write), db: AsyncSession = Depends(get_db)):
    loop = FiberLoop(**body.model_dump())
    db.add(loop)
    await db.commit()
    await db.refresh(loop)
    return loop


@router.put("/loops/{loop_id}", response_model=FiberLoopOut)
async def update_loop(loop_id: int, body: FiberLoopUpdate, user: User = Depends(require_write), db: AsyncSession = Depends(get_db)):
    loop = await db.get(FiberLoop, loop_id)
    if loop is None:
        raise HTTPException(status_code=404, detail="Loop not found")
    for field, value in body.model_dump(exclude_unset=True).items():
        setattr(loop, field, value)
    await db.commit()
    await db.refresh(loop)
    return loop


@router.delete("/loops/{loop_id}", status_code=204)
async def delete_loop(loop_id: int, user: User = Depends(require_write), db: AsyncSession = Depends(get_db)):
    loop = await db.get(FiberLoop, loop_id)
    if loop is None:
        raise HTTPException(status_code=404, detail="Loop not found")
    await db.delete(loop)
    await db.commit()


# ------------------------------------------------------------------ Cable Cuts
@router.get("/cuts", response_model=list[CableCutOut])
async def list_cuts(cable_id: int | None = None, status: str | None = None, db: AsyncSession = Depends(get_db)):
    q = select(CableCut).order_by(CableCut.id.desc())
    if cable_id is not None:
        q = q.where(CableCut.cable_id == cable_id)
    if status is not None:
        q = q.where(CableCut.status == status)
    res = await db.execute(q)
    cuts = res.scalars().all()
    result = []
    for cut in cuts:
        cut.splice_tj_name = ""
        if cut.splice_tj_id:
            tj = await db.get(TjBox, cut.splice_tj_id)
            if tj:
                cut.splice_tj_name = tj.name
        result.append(cut)
    return result


@router.post("/cuts", response_model=CableCutOut)
async def create_cut(body: CableCutCreate, user: User = Depends(require_write), db: AsyncSession = Depends(get_db)):
    cut = CableCut(**body.model_dump())
    db.add(cut)
    await db.commit()
    await db.refresh(cut)
    return cut


@router.put("/cuts/{cut_id}", response_model=CableCutOut)
async def update_cut(cut_id: int, body: CableCutUpdate, user: User = Depends(require_write), db: AsyncSession = Depends(get_db)):
    cut = await db.get(CableCut, cut_id)
    if cut is None:
        raise HTTPException(status_code=404, detail="Cable cut not found")
    for field, value in body.model_dump(exclude_unset=True).items():
        setattr(cut, field, value)
    await db.commit()
    await db.refresh(cut)
    cut.splice_tj_name = ""
    if cut.splice_tj_id:
        tj = await db.get(TjBox, cut.splice_tj_id)
        if tj:
            cut.splice_tj_name = tj.name
    return cut


@router.delete("/cuts/{cut_id}", status_code=204)
async def delete_cut(cut_id: int, user: User = Depends(require_write), db: AsyncSession = Depends(get_db)):
    cut = await db.get(CableCut, cut_id)
    if cut is None:
        raise HTTPException(status_code=404, detail="Cable cut not found")
    await db.delete(cut)
    await db.commit()


# Fiber core color sequence (standard 12-color)
CORE_COLOR_NAMES = [
    "Blue", "Orange", "Green", "Brown", "Slate", "White",
    "Red", "Black", "Yellow", "Violet", "Rose", "Aqua",
]

SUPPORTED_TJ_CAPACITIES = [4, 8, 10]


def _select_tj_capacity(core_count: int) -> int:
    """Select the smallest TJ capacity that can accommodate the required connections."""
    for cap in SUPPORTED_TJ_CAPACITIES:
        if cap >= core_count:
            return cap
    raise HTTPException(400, f"Core count {core_count} exceeds maximum supported TJ capacity ({max(SUPPORTED_TJ_CAPACITIES)}). Manual recovery required.")


@router.post("/cuts/{cut_id}/recover", response_model=CutRecoveryResult)
async def recover_cut(cut_id: int, user: User = Depends(require_write), db: AsyncSession = Depends(get_db)):
    """Automatically recover a cable cut: create TJ, connect cables, splice same-color cores."""
    from datetime import datetime, timezone

    # 1. Look up the cut
    cut = await db.get(CableCut, cut_id)
    if cut is None:
        raise HTTPException(404, "Cable cut not found")
    if cut.status == "repaired":
        raise HTTPException(400, "This cut is already repaired")

    # 2. Look up the cable
    cable = await db.get(Cable, cut.cable_id)
    if cable is None:
        raise HTTPException(400, "Associated cable not found")

    core_count = cable.core_count
    if core_count <= 0:
        raise HTTPException(400, "Cable has no cores to splice")

    # 3. Determine TJ capacity
    tj_capacity = _select_tj_capacity(core_count)

    # 4. Generate unique TJ ID
    res = await db.execute(select(TjBox.unique_id).where(TjBox.unique_id.like("TJ-%")).order_by(TjBox.id.desc()).limit(1))
    last = res.scalar()
    next_num = 5001
    if last:
        try:
            next_num = int(last.split("-")[1]) + 1
        except (IndexError, ValueError):
            pass
    tj_unique_id = f"TJ-{next_num:04d}"

    # 5. Create new TJ at cut location
    new_tj = TjBox(
        unique_id=tj_unique_id,
        name=f"Recovery {tj_unique_id}",
        box_type="regular_tj",
        tj_port=tj_capacity,
        capacity=tj_capacity,
        tray_count=1,
        splice_per_tray=tj_capacity,
        lat=cut.lat,
        lng=cut.lng,
        address="",
        notes=f"Auto-created for cut recovery on {cable.code}",
    )
    db.add(new_tj)
    await db.flush()  # Get the new TJ's ID

    # 6. Create splices for each core (same-color matching)
    splices_created = 0
    splice_details = []
    unmatched_cores = []

    for core_idx in range(1, core_count + 1):
        color_name = CORE_COLOR_NAMES[(core_idx - 1) % len(CORE_COLOR_NAMES)]

        # Check for duplicate splice (idempotency)
        existing = (await db.execute(
            select(Splice).where(
                Splice.tj_id == new_tj.id,
                Splice.status.in_(["active", "spare"]),
                Splice.cable_a_id == cable.id,
                Splice.core_a == core_idx,
                Splice.cable_b_id == cable.id,
                Splice.core_b == core_idx,
            )
        )).scalars().first()

        if existing:
            # Already spliced (idempotent)
            splices_created += 1
            splice_details.append(CutRecoverySplice(
                core_index=core_idx, color=color_name,
                cable_a_id=cable.id, cable_b_id=cable.id,
            ))
            continue

        # Create splice: same cable, same core (pass-through at recovery TJ)
        splice = Splice(
            tj_id=new_tj.id,
            cable_a_id=cable.id,
            core_a=core_idx,
            cable_b_id=cable.id,
            core_b=core_idx,
            tray_id=1,
            status="active",
            notes=f"Auto-spliced: {color_name} core {core_idx}",
        )
        db.add(splice)
        splices_created += 1
        splice_details.append(CutRecoverySplice(
            core_index=core_idx, color=color_name,
            cable_a_id=cable.id, cable_b_id=cable.id,
        ))

    # 7. Update cut status
    cut.status = "repaired"
    cut.repair_date = datetime.now(timezone.utc)
    cut.splice_tj_id = new_tj.id
    cut.notes = (cut.notes + "\n" if cut.notes else "") + f"Auto-recovered: {tj_unique_id} with {splices_created} splices"

    await db.commit()
    await db.refresh(new_tj)

    return CutRecoveryResult(
        tj_id=new_tj.id,
        tj_unique_id=new_tj.unique_id,
        tj_name=new_tj.name,
        tj_capacity=tj_capacity,
        cable_id=cable.id,
        cable_code=cable.code,
        core_count=core_count,
        splices_created=splices_created,
        splices=splice_details,
        unmatched_cores=unmatched_cores,
    )


# ------------------------------------------------------------ Export / Import
@router.get("/export")
async def export_fiber(db: AsyncSession = Depends(get_db)):
    import openpyxl
    wb = openpyxl.Workbook()

    # TJ Boxes sheet
    ws = wb.active
    ws.title = "TJ Boxes"
    ws.append(["Name", "Port", "Latitude", "Longitude", "Address", "Note"])
    res = await db.execute(select(TjBox).order_by(TjBox.id))
    for t in res.scalars().all():
        ws.append([t.name, t.tj_port, t.lat, t.lng, t.address, t.notes])

    # Splitters sheet
    ws2 = wb.create_sheet("Splitters")
    ws2.append(["Name", "Split Ratio", "TJ Box Name", "Input Core", "Output Cores", "Latitude", "Longitude", "Notes"])
    res = await db.execute(select(Splitter).order_by(Splitter.id))
    for s in res.scalars().all():
        box_name = ""
        if s.tj_box_id:
            box = await db.get(TjBox, s.tj_box_id)
            box_name = box.name if box else ""
        ws2.append([s.name, s.split_ratio, box_name, s.input_core, s.output_cores, s.lat, s.lng, s.notes])

    # Cables sheet
    ws3 = wb.create_sheet("Cables")
    ws3.append(["Link ID", "Link Name", "Code", "Core Count", "Type", "Manufacturer", "Year", "Route Type", "Notes", "Segment Lat", "Segment Lng", "Segment End Lat", "Segment End Lng"])
    res = await db.execute(select(Cable).order_by(Cable.id))
    for c in res.scalars().all():
        segs = (await db.execute(select(CableSegment).where(CableSegment.cable_id == c.id).order_by(CableSegment.order_index))).scalars().all()
        if segs:
            for seg in segs:
                ws3.append([c.link_id, c.link_name, c.code, c.core_count, c.cable_type, c.manufacturer, c.manufacturing_year, c.route_type, c.notes,
                            seg.start_lat, seg.start_lng, seg.end_lat, seg.end_lng])
        else:
            ws3.append([c.link_id, c.link_name, c.code, c.core_count, c.cable_type, c.manufacturer, c.manufacturing_year, c.route_type, c.notes, "", "", "", ""])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=fiber_network.xlsx"})


@router.post("/import")
async def import_fiber(file: UploadFile = File(...), user: User = Depends(require_write), db: AsyncSession = Depends(get_db)):
    import openpyxl
    content = await file.read()
    wb = openpyxl.load_workbook(io.BytesIO(content))
    imported = {"tj_boxes": 0, "splitters": 0, "cables": 0}

    # Find next IDs
    res = await db.execute(select(TjBox.unique_id).where(TjBox.unique_id.like("TJ-%")).order_by(TjBox.id.desc()).limit(1))
    last = res.scalar()
    tj_next = 5001
    if last:
        try: tj_next = int(last.split("-")[1]) + 1
        except: pass

    res = await db.execute(select(Splitter.unique_id).where(Splitter.unique_id.like("SP-%")).order_by(Splitter.id.desc()).limit(1))
    last = res.scalar()
    sp_next = 1001
    if last:
        try: sp_next = int(last.split("-")[1]) + 1
        except: pass

    def _map_columns(ws):
        """Build {header_name: col_index} dict from row 1, then return list-of-dicts for data rows."""
        headers = [str(c.value or "").strip().lower() for c in ws[1]] if ws.max_row else []
        rows = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            rows.append({headers[i]: row[i] for i in range(min(len(headers), len(row)))})
        return rows

    # TJ Boxes
    if "TJ Boxes" in wb.sheetnames:
        tj_rows = _map_columns(wb["TJ Boxes"])
        for r in tj_rows:
            name = str(r.get("name") or "").strip()
            if not name:
                continue
            box = TjBox(
                unique_id=f"TJ-{tj_next}",
                name=name, box_type="regular_tj",
                tj_port=int(r.get("port") or 8),
                splice_per_tray=12, tray_count=1, capacity=12,
                lat=float(r.get("latitude") or 0), lng=float(r.get("longitude") or 0),
                address=str(r.get("address") or ""), notes=str(r.get("note") or ""),
            )
            db.add(box)
            tj_next += 1
            imported["tj_boxes"] += 1

    # Splitters
    if "Splitters" in wb.sheetnames:
        sp_rows = _map_columns(wb["Splitters"])
        for r in sp_rows:
            name = str(r.get("name") or "").strip()
            lat = r.get("latitude")
            lng = r.get("longitude")
            if not name or not lat or not lng:
                continue
            tj_box_id = None
            tj_name = str(r.get("tj box name") or r.get("tj_box_name") or "").strip()
            if tj_name:
                box_res = await db.execute(select(TjBox).where(TjBox.name == tj_name))
                box = box_res.scalars().first()
                if box:
                    tj_box_id = box.id
            splitter = Splitter(
                unique_id=f"SP-{sp_next}",
                name=name, split_ratio=int(r.get("split ratio") or 2),
                tj_box_id=tj_box_id, input_core=int(r.get("input core") or 0),
                output_cores=str(r.get("output cores") or ""),
                lat=float(lat), lng=float(lng),
                notes=str(r.get("notes") or r.get("note") or ""),
            )
            db.add(splitter)
            sp_next += 1
            imported["splitters"] += 1

    # Cables
    if "Cables" in wb.sheetnames:
        cable_cache: dict[str, Cable] = {}
        cbl_rows = _map_columns(wb["Cables"])
        for r in cbl_rows:
            link_id = str(r.get("link id") or "").strip()
            if not link_id:
                max_res = await db.execute(select(Cable.id).order_by(Cable.id.desc()).limit(1))
                max_id = max_res.scalar() or 0
                link_id = f"LINK-{max_id + 1001}"
            if link_id not in cable_cache:
                cable = Cable(
                    link_id=link_id,
                    link_name=str(r.get("link name") or ""), code=str(r.get("code") or ""),
                    core_count=int(r.get("core count") or 12), cable_type=str(r.get("type") or "round"),
                    manufacturer=str(r.get("manufacturer") or ""), manufacturing_year=int(r.get("year") or 0),
                    route_type=str(r.get("route type") or "driving"), notes=str(r.get("notes") or r.get("note") or ""),
                )
                db.add(cable)
                await db.flush()
                cable_cache[link_id] = cable
                imported["cables"] += 1
            seg_lat = r.get("segment lat")
            seg_lng = r.get("segment lng")
            seg_end_lat = r.get("segment end lat")
            seg_end_lng = r.get("segment end lng")
            if seg_lat and seg_lng and seg_end_lat and seg_end_lng:
                seg = CableSegment(
                    cable_id=cable_cache[link_id].id,
                    start_lat=float(seg_lat), start_lng=float(seg_lng),
                    end_lat=float(seg_end_lat), end_lng=float(seg_end_lng),
                )
                db.add(seg)

    await db.commit()
    return imported


# ──────────────────────────────────────────────────────────────── splices

@router.get("/splices", response_model=list[SpliceOut])
async def list_splices(tj_id: int | None = None, limit: int = 200, offset: int = 0, db: AsyncSession = Depends(get_db)):
    q = select(Splice)
    if tj_id is not None:
        q = q.where(Splice.tj_id == tj_id)
    q = q.order_by(Splice.tray_id, Splice.id).limit(limit).offset(offset)
    res = await db.execute(q)
    splices = res.scalars().all()
    # Batch lookup
    cable_ids = set()
    splitter_ids = set()
    for sp in splices:
        if sp.cable_a_id: cable_ids.add(sp.cable_a_id)
        if sp.cable_b_id: cable_ids.add(sp.cable_b_id)
        if sp.splitter_a_id: splitter_ids.add(sp.splitter_a_id)
        if sp.splitter_b_id: splitter_ids.add(sp.splitter_b_id)
    cable_map = {}
    if cable_ids:
        cables_res = await db.execute(select(Cable.id, Cable.code).where(Cable.id.in_(cable_ids)))
        cable_map = {c.id: c.code for c in cables_res.all()}
    splitter_map = {}
    splitter_ratio_map = {}
    if splitter_ids:
        splitters_res = await db.execute(select(Splitter.id, Splitter.name, Splitter.split_ratio).where(Splitter.id.in_(splitter_ids)))
        for s in splitters_res.all():
            splitter_map[s.id] = s.name
            splitter_ratio_map[s.id] = s.split_ratio
    out = []
    for sp in splices:
        d = SpliceOut.model_validate(sp)
        d.cable_a_code = cable_map.get(sp.cable_a_id, "") if sp.cable_a_id else ""
        d.cable_b_code = cable_map.get(sp.cable_b_id, "") if sp.cable_b_id else ""
        d.splitter_a_name = splitter_map.get(sp.splitter_a_id, "") if sp.splitter_a_id else ""
        d.splitter_b_name = splitter_map.get(sp.splitter_b_id, "") if sp.splitter_b_id else ""
        d.splitter_a_ratio = splitter_ratio_map.get(sp.splitter_a_id, 0) if sp.splitter_a_id else 0
        d.splitter_b_ratio = splitter_ratio_map.get(sp.splitter_b_id, 0) if sp.splitter_b_id else 0
        out.append(d)
    return out


@router.post("/splices", response_model=SpliceOut, status_code=201, dependencies=[Depends(require_write)])
async def create_splice(body: SpliceCreate, db: AsyncSession = Depends(get_db)):
    from sqlalchemy import and_, or_
    
    # Validate: must have at least one endpoint (cable or splitter on each side)
    if not body.cable_a_id and not body.splitter_a_id:
        raise HTTPException(400, "Must specify either cable_a_id or splitter_a_id")
    if not body.cable_b_id and not body.splitter_b_id:
        raise HTTPException(400, "Must specify either cable_b_id or splitter_b_id")
    
    # Validate: cannot have both cable and splitter on same side
    if body.cable_a_id and body.splitter_a_id:
        raise HTTPException(400, "Cannot specify both cable_a_id and splitter_a_id")
    if body.cable_b_id and body.splitter_b_id:
        raise HTTPException(400, "Cannot specify both cable_b_id and splitter_b_id")
    
    # Validate: cannot self-splice (same cable+core or same splitter+port)
    if body.cable_a_id and body.cable_b_id and body.cable_a_id == body.cable_b_id and body.core_a == body.core_b:
        raise HTTPException(400, "Cannot splice a cable core to itself")
    if body.splitter_a_id and body.splitter_b_id and body.splitter_a_id == body.splitter_b_id and body.port_a == body.port_b:
        raise HTTPException(400, "Cannot splice a splitter port to itself")
    
    # Validate: both endpoints must belong to the same TJ
    if body.cable_a_id:
        cable_a = await db.get(Cable, body.cable_a_id)
        if not cable_a:
            raise HTTPException(400, "Cable A not found")
        if cable_a.src_tj_id != body.tj_id and cable_a.dst_tj_id != body.tj_id:
            raise HTTPException(400, "Cable A is not connected to this TJ")
    if body.cable_b_id:
        cable_b = await db.get(Cable, body.cable_b_id)
        if not cable_b:
            raise HTTPException(400, "Cable B not found")
        if cable_b.src_tj_id != body.tj_id and cable_b.dst_tj_id != body.tj_id:
            raise HTTPException(400, "Cable B is not connected to this TJ")
    if body.splitter_a_id:
        splitter_a = await db.get(Splitter, body.splitter_a_id)
        if not splitter_a:
            raise HTTPException(400, "Splitter A not found")
        if splitter_a.tj_box_id != body.tj_id:
            raise HTTPException(400, "Splitter A is not hosted at this TJ")
    if body.splitter_b_id:
        splitter_b = await db.get(Splitter, body.splitter_b_id)
        if not splitter_b:
            raise HTTPException(400, "Splitter B not found")
        if splitter_b.tj_box_id != body.tj_id:
            raise HTTPException(400, "Splitter B is not hosted at this TJ")
    
    # Check for duplicate splice
    existing_dup = (await db.execute(
        select(Splice).where(
            Splice.tj_id == body.tj_id,
            Splice.status.in_(["active", "spare"]),
            or_(
                and_(
                    Splice.cable_a_id == body.cable_a_id, Splice.core_a == body.core_a,
                    Splice.cable_b_id == body.cable_b_id, Splice.core_b == body.core_b,
                    Splice.splitter_a_id == body.splitter_a_id, Splice.port_a == body.port_a,
                    Splice.splitter_b_id == body.splitter_b_id, Splice.port_b == body.port_b,
                ),
                and_(
                    Splice.cable_a_id == body.cable_b_id, Splice.core_a == body.core_b,
                    Splice.cable_b_id == body.cable_a_id, Splice.core_b == body.core_a,
                    Splice.splitter_a_id == body.splitter_b_id, Splice.port_a == body.port_b,
                    Splice.splitter_b_id == body.splitter_a_id, Splice.port_b == body.port_a,
                ),
            )
        )
    )).scalars().all()
    if existing_dup:
        raise HTTPException(400, "This splice already exists")
    
    # Validate core/port occupancy
    async def _check_occupied(cable_id, core, splitter_id, port, side_label):
        if cable_id:
            occ = (await db.execute(
                select(Splice).where(
                    Splice.tj_id == body.tj_id,
                    Splice.status.in_(["active", "spare"]),
                    or_(
                        and_(Splice.cable_a_id == cable_id, Splice.core_a == core),
                        and_(Splice.cable_b_id == cable_id, Splice.core_b == core),
                    )
                )
            )).scalars().first()
            if occ:
                raise HTTPException(400, f"{side_label}: Core {core} is already occupied by splice #{occ.id}")
        elif splitter_id:
            occ = (await db.execute(
                select(Splice).where(
                    Splice.tj_id == body.tj_id,
                    Splice.status.in_(["active", "spare"]),
                    or_(
                        and_(Splice.splitter_a_id == splitter_id, Splice.port_a == port),
                        and_(Splice.splitter_b_id == splitter_id, Splice.port_b == port),
                    )
                )
            )).scalars().first()
            if occ:
                raise HTTPException(400, f"{side_label}: Port {port} is already occupied by splice #{occ.id}")

    await _check_occupied(body.cable_a_id, body.core_a, body.splitter_a_id, body.port_a, "Endpoint A")
    await _check_occupied(body.cable_b_id, body.core_b, body.splitter_b_id, body.port_b, "Endpoint B")

    splice = Splice(**body.model_dump())
    db.add(splice)
    await db.commit()
    await db.refresh(splice)
    
    # Build response
    d = SpliceOut.model_validate(splice)
    if splice.cable_a_id:
        ca = await db.get(Cable, splice.cable_a_id)
        d.cable_a_code = ca.code if ca else ""
    if splice.cable_b_id:
        cb = await db.get(Cable, splice.cable_b_id)
        d.cable_b_code = cb.code if cb else ""
    if splice.splitter_a_id:
        sa = await db.get(Splitter, splice.splitter_a_id)
        d.splitter_a_name = sa.name if sa else ""
        d.splitter_a_ratio = sa.split_ratio if sa else 0
    if splice.splitter_b_id:
        sb = await db.get(Splitter, splice.splitter_b_id)
        d.splitter_b_name = sb.name if sb else ""
        d.splitter_b_ratio = sb.split_ratio if sb else 0
    return d


@router.put("/splices/{splice_id}", response_model=SpliceOut, dependencies=[Depends(require_write)])
async def update_splice(splice_id: int, body: SpliceUpdate, db: AsyncSession = Depends(get_db)):
    res = await db.execute(select(Splice).where(Splice.id == splice_id))
    splice = res.scalar_one_or_none()
    if not splice:
        raise HTTPException(404, "Splice not found")

    # Determine effective values after update
    eff = body.model_dump(exclude_unset=True)
    eff_cable_a = eff.get("cable_a_id", splice.cable_a_id)
    eff_core_a = eff.get("core_a", splice.core_a)
    eff_cable_b = eff.get("cable_b_id", splice.cable_b_id)
    eff_core_b = eff.get("core_b", splice.core_b)
    eff_splitter_a = eff.get("splitter_a_id", splice.splitter_a_id)
    eff_splitter_b = eff.get("splitter_b_id", splice.splitter_b_id)
    eff_port_a = eff.get("port_a", splice.port_a)
    eff_port_b = eff.get("port_b", splice.port_b)
    eff_status = eff.get("status", splice.status)

    # If setting to active/spare, validate cores/ports are not occupied
    if eff_status in ("active", "spare"):
        from sqlalchemy import and_, or_
        
        async def _check_occupied_for_update(cable_id, core, splitter_id, port, side_label):
            if cable_id:
                occ = (await db.execute(
                    select(Splice).where(
                        Splice.tj_id == splice.tj_id,
                        Splice.id != splice_id,
                        Splice.status.in_(["active", "spare"]),
                        or_(
                            and_(Splice.cable_a_id == cable_id, Splice.core_a == core),
                            and_(Splice.cable_b_id == cable_id, Splice.core_b == core),
                        )
                    )
                )).scalars().first()
                if occ:
                    raise HTTPException(400, f"{side_label}: Core {core} is already occupied by splice #{occ.id}")
            elif splitter_id:
                occ = (await db.execute(
                    select(Splice).where(
                        Splice.tj_id == splice.tj_id,
                        Splice.id != splice_id,
                        Splice.status.in_(["active", "spare"]),
                        or_(
                            and_(Splice.splitter_a_id == splitter_id, Splice.port_a == port),
                            and_(Splice.splitter_b_id == splitter_id, Splice.port_b == port),
                        )
                    )
                )).scalars().first()
                if occ:
                    raise HTTPException(400, f"{side_label}: Port {port} is already occupied by splice #{occ.id}")

        await _check_occupied_for_update(eff_cable_a, eff_core_a, eff_splitter_a, eff_port_a, "Endpoint A")
        await _check_occupied_for_update(eff_cable_b, eff_core_b, eff_splitter_b, eff_port_b, "Endpoint B")

    for k, v in eff.items():
        setattr(splice, k, v)
    await db.commit()
    await db.refresh(splice)
    
    # Build response
    d = SpliceOut.model_validate(splice)
    if splice.cable_a_id:
        ca = await db.get(Cable, splice.cable_a_id)
        d.cable_a_code = ca.code if ca else ""
    if splice.cable_b_id:
        cb = await db.get(Cable, splice.cable_b_id)
        d.cable_b_code = cb.code if cb else ""
    if splice.splitter_a_id:
        sa = await db.get(Splitter, splice.splitter_a_id)
        d.splitter_a_name = sa.name if sa else ""
        d.splitter_a_ratio = sa.split_ratio if sa else 0
    if splice.splitter_b_id:
        sb = await db.get(Splitter, splice.splitter_b_id)
        d.splitter_b_name = sb.name if sb else ""
        d.splitter_b_ratio = sb.split_ratio if sb else 0
    return d


@router.delete("/splices/{splice_id}", status_code=204, dependencies=[Depends(require_write)])
async def delete_splice(splice_id: int, db: AsyncSession = Depends(get_db)):
    res = await db.execute(select(Splice).where(Splice.id == splice_id))
    splice = res.scalar_one_or_none()
    if not splice:
        raise HTTPException(404, "Splice not found")
    await db.delete(splice)
    await db.commit()


@router.get("/splices/unused-cores")
async def unused_cores(tj_id: int, db: AsyncSession = Depends(get_db)):
    """Return unused (spare) cores for each cable connected to a TJ."""
    tj_res = await db.execute(select(TjBox).where(TjBox.id == tj_id))
    tj = tj_res.scalar_one_or_none()
    if not tj:
        raise HTTPException(404, "TJ not found")
    # Find cable IDs connected to this TJ via segments (filter in query)
    from sqlalchemy import or_
    segs_res = await db.execute(
        select(CableSegment.cable_id).where(
            or_(
                ( CableSegment.start_lat.between(tj.lat - 0.001, tj.lat + 0.001)) & (CableSegment.start_lng.between(tj.lng - 0.001, tj.lng + 0.001)),
                (CableSegment.end_lat.between(tj.lat - 0.001, tj.lat + 0.001)) & (CableSegment.end_lng.between(tj.lng - 0.001, tj.lng + 0.001)),
            )
        ).distinct()
    )
    cable_ids = set(segs_res.scalars().all())
    if not cable_ids:
        return []
    # Get all active/spare splices at this TJ (broken splices don't occupy cores)
    splices_res = await db.execute(
        select(Splice.cable_a_id, Splice.core_a, Splice.cable_b_id, Splice.core_b)
        .where(Splice.tj_id == tj_id, Splice.status.in_(["active", "spare"]))
    )
    used = set()
    for row in splices_res.all():
        used.add((row.cable_a_id, row.core_a))
        used.add((row.cable_b_id, row.core_b))
    # Also mark cores allocated to hosted splitters as occupied
    splitter_res = await db.execute(
        select(Splitter).where(Splitter.tj_box_id == tj_id)
    )
    for sp in splitter_res.scalars().all():
        if sp.input_core:
            # Find which cable connects to this TJ to assign the input core
            for cid in cable_ids:
                used.add((cid, sp.input_core))
        if sp.output_cores:
            for part in sp.output_cores.split(","):
                part = part.strip()
                if part.isdigit():
                    for cid in cable_ids:
                        used.add((cid, int(part)))
    # Get cable info
    cables_res = await db.execute(select(Cable).where(Cable.id.in_(cable_ids)))
    cables = cables_res.scalars().all()
    result = []
    for c in cables:
        spare = [i for i in range(1, c.core_count + 1) if (c.id, i) not in used]
        occupied = [i for i in range(1, c.core_count + 1) if (c.id, i) in used]
        result.append({"cable_id": c.id, "cable_code": c.code, "core_count": c.core_count, "spare_cores": spare, "occupied_cores": occupied})
    return result


@router.get("/noc-pop-map")
async def get_noc_pop_map(db: AsyncSession = Depends(get_db)):
    from ..models import Noc, Pop, OLTDevice
    nocs = (await db.execute(select(Noc))).scalars().all()
    pops = (await db.execute(select(Pop))).scalars().all()
    olts = (await db.execute(select(OLTDevice))).scalars().all()

    noc_items = []
    for n in nocs:
        devices = [o for o in olts if o.noc_id == n.id]
        noc_items.append({
            "id": n.id, "name": n.name, "type": "noc",
            "lat": n.gps_lat, "lng": n.gps_lng, "address": n.address,
            "contact_name": n.contact_name, "contact_phone": n.contact_phone,
            "device_count": len(devices),
            "devices": [{"id": d.id, "name": d.name, "ip": d.ip, "status": d.status, "pon_type": d.pon_type} for d in devices],
        })

    pop_items = []
    for p in pops:
        devices = [o for o in olts if o.pop_id == p.id]
        pop_items.append({
            "id": p.id, "name": p.name, "type": "pop",
            "lat": p.gps_lat, "lng": p.gps_lng, "address": p.address,
            "contact_name": p.contact_name, "contact_phone": p.contact_phone,
            "device_count": len(devices),
            "devices": [{"id": d.id, "name": d.name, "ip": d.ip, "status": d.status, "pon_type": d.pon_type} for d in devices],
        })

    return {"nocs": noc_items, "pops": pop_items}
